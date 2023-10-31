# -*-coding:Utf-8 -*


# For Excel
from openpyxl import load_workbook

# For logs
import random
import logging
import os
import sys

# To get line number for logs
from inspect import currentframe, getframeinfo
import inspect

# Dates
import datetime
import time

# For args
import argparse

import getopt

# Regex
import re

# param
end_line_character_in_text_file = "\n"
logger_level = logging.INFO

# wordContainingUnderscoreToTranformRegexCompiledPattern = re.compile(wordContainingUnderscoreToTranformRegexPatternAsString)


def printAndLogCriticalAndKill(toPrintAndLog):
    log_timestamp = time.asctime(time.localtime(time.time()))

    previous_stack = inspect.stack(1)[1]
    line_number = previous_stack.lineno

    print(log_timestamp + '\t' + "line#" +
          str(line_number) + '\t' + toPrintAndLog)
    logging.critical("line#" + str(line_number) + '\t' + toPrintAndLog)
    sys.exit()


def printAndLogInfo(toPrintAndLog):

    log_timestamp = time.asctime(time.localtime(time.time()))

    previous_stack = inspect.stack(1)[1]
    line_number = previous_stack.lineno

    print(log_timestamp + '\t' + "line#" +
          str(line_number) + '\t' + str() + toPrintAndLog)
    logging.info("line#" + str(line_number) + '\t' + toPrintAndLog)


def printAndLogWarning(toPrintAndLog):
    log_timestamp = time.asctime(time.localtime(time.time()))

    previous_stack = inspect.stack(1)[1]
    line_number = previous_stack.lineno

    print(log_timestamp + '\t' + "line#" +
          str(line_number) + '\t' + toPrintAndLog)
    logging.warning("line#" + str(line_number) + '\t' + toPrintAndLog)


def printAndLogError(toPrintAndLog):
    log_timestamp = time.asctime(time.localtime(time.time()))

    previous_stack = inspect.stack(1)[1]
    line_number = previous_stack.lineno

    print(log_timestamp + '\t' + "!!ERROR!!")
    print(log_timestamp + '\t' "line#" + str(line_number))
    print(log_timestamp + '\t' + toPrintAndLog)
    logging.error("line#" + str(line_number) + '\t' + toPrintAndLog)


def configureLogger(log_file_name):
    logger_directory = "logs"

    if not os.path.exists(logger_directory):
        os.makedirs(logger_directory)

    print(time.asctime(time.localtime(time.time())) +
          '\t' + "Logger level:" + str(logger_level))

    logging.basicConfig(level=logger_level,
                        format='%(asctime)s %(levelname)-8s %(message)s',
                        datefmt='%a, %d %b %Y %H:%M:%S',
                        filename=logger_directory + '\\' + log_file_name,
                        filemode='w')
    # logging.debug
    # logging.info
    # logging.warning
    # logging.error
    # logging.critical


def getAllUndescoreContainingTextWordsInFile(fileReadContent):

    allUndescoreContainingTextWordsAsSet = set()

    fileReadContent = fileReadContent.replace(
        end_line_character_in_text_file, " ")
    fileReadContent = fileReadContent.replace(".", " ")
    fileReadContent = fileReadContent.replace("  ", " ")
    fileReadContent = fileReadContent.replace("\t", " ")

    fileReadContent = fileReadContent.replace("(", " ")
    fileReadContent = fileReadContent.replace(")", " ")
    fileReadContent = fileReadContent.replace(";", " ")

    
    fileReadContent = fileReadContent.replace(":", " ")
    
    fileReadContent = fileReadContent.replace(",", " ")

    allCodeSyntaxElementsAsList = fileReadContent.split(" ")
    allCodeSyntaxElementsAsSet = set(allCodeSyntaxElementsAsList)
    for codeSyntaxElement in allCodeSyntaxElementsAsSet:
        if "_" in codeSyntaxElement:
            allUndescoreContainingTextWordsAsSet.add(codeSyntaxElement)

    return allUndescoreContainingTextWordsAsSet



def getUnderscoreContainingCodeWordTransformedIntoCamelCase(initialWordWithUnderscore):
    codeWordTransformedIntoCamelCase = ""
    nextLetterMustBeCapitalLetter = False

    for letter in initialWordWithUnderscore:
        if letter == "_":
            nextLetterMustBeCapitalLetter = True
        elif nextLetterMustBeCapitalLetter:
            codeWordTransformedIntoCamelCase += letter.upper()
            nextLetterMustBeCapitalLetter = False
        else:
            codeWordTransformedIntoCamelCase += letter

    printAndLogInfo(initialWordWithUnderscore +
                    " transformed to camel case is:" + codeWordTransformedIntoCamelCase)
    return codeWordTransformedIntoCamelCase


def replaceAllTextByCamelCaseInFile(fileFullPath, path, fileNameWithoutExtension, fileExtension, noOperation):
    camelCaseFileContent = None

    # Opening the file in read and write mode
    with open(fileFullPath, 'r+') as f:
        # Reading the file data and store
        # it in a file variable
        fileReadContent = f.read()
        camelCaseFileContent = fileReadContent

        allUndescoreContainingTextWordsInFileAsSet = getAllUndescoreContainingTextWordsInFile(
            camelCaseFileContent)

        for undescoreContainingTextWord in allUndescoreContainingTextWordsInFileAsSet:
            allowed = False
            for allowedRegexCompiledPattern in allowedRegexCompiledPatternsList:
                
                if undescoreContainingTextWord == "EXTERNAL_FRAME_WIDTH":
                    pause = 1
                if allowedRegexCompiledPattern.match(undescoreContainingTextWord):
                    logging.info(undescoreContainingTextWord +  " is allowed thanks to for regex: " + allowedRegexCompiledPattern.pattern)

                    allowed = True
                else:
                    logging.info(undescoreContainingTextWord +  " is not allowed for regex: " + allowedRegexCompiledPattern.pattern)

            if not allowed:
                codeWordTransformedIntoCamelCase = getUnderscoreContainingCodeWordTransformedIntoCamelCase(
                    undescoreContainingTextWord)
                camelCaseFileContent = camelCaseFileContent.replace(
                    undescoreContainingTextWord, codeWordTransformedIntoCamelCase)

    if not noOperation:
        # Opening our text file in write only
        # mode to write the replaced content
        with open(fileFullPath, 'w') as file:

            # Writing the replaced data in our
            # text file
            file.write(camelCaseFileContent)
    else:
        logging.debug("Output file would be:" + camelCaseFileContent)

def transformTextToValidExcelDate(initialText):
    transformedText = initialText.replace("UTC+1","")
    transformedText = transformedText.replace("UTC+2","")
    transformedText = transformedText.replace("à ","")
    return transformedText

def transformTextInColumnToValidExcelDate(sheet, columnName):
    printAndLogInfo("transformTextInColumnToValidExcelDate for column:" + columnName)

    submitDateColumnCells = sheet[columnName]
    for submitDateColumnCell in submitDateColumnCells:
        initialText = submitDateColumnCell.value
        newText = transformTextToValidExcelDate(initialText)
        cellCoordinate = submitDateColumnCell.coordinate
        sheetCell = sheet[cellCoordinate]
        sheet[cellCoordinate]  = newText
        submitDateColumnCell = newText


def UpdateQueryResultToMatchCMCExpectedFormatNTOAnalyseCFX(argv):

    workbook = load_workbook(filename="QueryResult.xlsx")
    workbook_sheetnames = workbook.sheetnames 
    print(workbook_sheetnames)

    sheet = workbook.active
    printAndLogInfo("sheet:" + str(sheet))

    #print(sheet)
    #sheet

    sheet_title = sheet.title
    printAndLogInfo("sheet title:"+ sheet_title)

    submitDateColumnName = 'G'
    transformTextInColumnToValidExcelDate(sheet, submitDateColumnName)
  
    historyOfLastActionHistoryActionTimestampColumnName = 'M'
    transformTextInColumnToValidExcelDate(sheet, historyOfLastActionHistoryActionTimestampColumnName)
       
    
    printAndLogInfo("Save workbook")
    workbook.save(filename="output.xlsx")



def main(argv):
    log_file_name = 'UpdateQueryResultToMatchCMCExpectedFormatNTOAnalyseCFX' + '_' + \
        str(random.randrange(100000)) + ".log"

    # log_file_name = 'TransformCodeToCamelCase' + "." +  str(random.randrange(10000)) + ".log"
    configureLogger(log_file_name)

    printAndLogInfo('Start application. Log file name: ' + log_file_name)
    UpdateQueryResultToMatchCMCExpectedFormatNTOAnalyseCFX(argv)
    printAndLogInfo('End application')


if __name__ == '__main__':
    main(sys.argv[1:])
