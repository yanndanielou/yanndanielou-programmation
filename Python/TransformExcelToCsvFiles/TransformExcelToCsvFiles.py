# -*-coding:Utf-8 -*

# parse folder
import glob

# For Excel to Csv
import pandas


# For Excel
from openpyxl import load_workbook

# For logs
import random
import logging
import os
import sys

import param;

# To get line number for logs
from inspect import currentframe, getframeinfo
import inspect

# Dates
import datetime
import time

# For args
import argparse

import getopt

# Regex
import re

# param
end_line_character_in_text_file = "\n"
logger_level = logging.INFO

# wordContainingUnderscoreToTranformRegexCompiledPattern = re.compile(wordContainingUnderscoreToTranformRegexPatternAsString)


def printAndLogCriticalAndKill(toPrintAndLog):
    log_timestamp = time.asctime(time.localtime(time.time()))

    previous_stack = inspect.stack(1)[1]
    line_number = previous_stack.lineno

    print(log_timestamp + '\t' + "line#" +
          str(line_number) + '\t' + toPrintAndLog)
    logging.critical("line#" + str(line_number) + '\t' + toPrintAndLog)
    sys.exit()


def printAndLogInfo(toPrintAndLog):

    log_timestamp = time.asctime(time.localtime(time.time()))

    previous_stack = inspect.stack(1)[1]
    line_number = previous_stack.lineno

    print(log_timestamp + '\t' + "line#" +
          str(line_number) + '\t' + str() + toPrintAndLog)
    logging.info("line#" + str(line_number) + '\t' + toPrintAndLog)


def printAndLogWarning(toPrintAndLog):
    log_timestamp = time.asctime(time.localtime(time.time()))

    previous_stack = inspect.stack(1)[1]
    line_number = previous_stack.lineno

    print(log_timestamp + '\t' + "line#" +
          str(line_number) + '\t' + toPrintAndLog)
    logging.warning("line#" + str(line_number) + '\t' + toPrintAndLog)


def printAndLogError(toPrintAndLog):
    log_timestamp = time.asctime(time.localtime(time.time()))

    previous_stack = inspect.stack(1)[1]
    line_number = previous_stack.lineno

    print(log_timestamp + '\t' + "!!ERROR!!")
    print(log_timestamp + '\t' "line#" + str(line_number))
    print(log_timestamp + '\t' + toPrintAndLog)
    logging.error("line#" + str(line_number) + '\t' + toPrintAndLog)


def configureLogger(log_file_name):
    logger_directory = "logs"

    if not os.path.exists(logger_directory):
        os.makedirs(logger_directory)

    print(time.asctime(time.localtime(time.time())) +
          '\t' + "Logger level:" + str(logger_level))

    logging.basicConfig(level=logger_level,
                        format='%(asctime)s %(levelname)-8s %(message)s',
                        datefmt='%a, %d %b %Y %H:%M:%S',
                        filename=logger_directory + '\\' + log_file_name,
                        filemode='w')
    # logging.debug
    # logging.info
    # logging.warning
    # logging.error
    # logging.critical



def transformTextToValidExcelDate(initialText):
    transformedText = initialText.replace("UTC+1","")
    transformedText = transformedText.replace("UTC+2","")
    transformedText = transformedText.replace("à ","")
    return transformedText


def getListOfExcelFilesNames():
    dir_path = r'*.xls*'
    res = glob.glob(dir_path)
    return res

def transformExcelFileToCsvFiles(excelFileNameWithExtension):

    printAndLogInfo("Load workbook:" + excelFileNameWithExtension)
    workbook = load_workbook(filename=excelFileNameWithExtension)

    excelFileNameWithoutExtension = excelFileNameWithExtension.split(".")[0]

    workbook_sheetnames = workbook.sheetnames 
    printAndLogInfo("Workbook:" + excelFileNameWithoutExtension + " has following sheets:" + str(workbook_sheetnames))
    
    if not os.path.exists(excelFileNameWithoutExtension):
        printAndLogInfo("Create folder " + excelFileNameWithoutExtension) 
        os.makedirs(excelFileNameWithoutExtension)

    for workbookSheetName in workbook_sheetnames:
        read_file = pandas.read_excel (excelFileNameWithExtension, sheet_name=workbookSheetName)
        output_csv_file = excelFileNameWithoutExtension + '/' + workbookSheetName + '.csv'
        printAndLogInfo("Extract sheet " + workbookSheetName + " to " + output_csv_file)
        read_file.to_csv (output_csv_file, sep=param.csv_separator, index = None, header=True)


def transformAllExcelFilesToCsvFiles(argv):

    listOfExcelFileNames = getListOfExcelFilesNames()

    printAndLogInfo("Number of excel files found:" + str(len(listOfExcelFileNames)) + " : " + str(listOfExcelFileNames))
 
    for excelFileName in listOfExcelFileNames:
        transformExcelFileToCsvFiles(excelFileName)
    
  


def main(argv):
    log_file_name = 'TransformExcelToCsvFiles' + '_' + \
        str(random.randrange(100000)) + ".log"

    # log_file_name = 'TransformCodeToCamelCase' + "." +  str(random.randrange(10000)) + ".log"
    configureLogger(log_file_name)

    printAndLogInfo('Start application. Log file name: ' + log_file_name)
    transformAllExcelFilesToCsvFiles(argv)
    printAndLogInfo('End application')


if __name__ == '__main__':
    main(sys.argv[1:])
